"""XLSX ingestion (PR13b-2), behind the same DatasetSpec as CSV/TSV.

A spreadsheet is not a delimited file with a different separator. It is typed,
formula-bearing and multi-sheet, and each of those is a way for a value to arrive different
from what a human saw in Excel. These tests are organised around the two questions that
matter:

* does the *same* contract hold — same spec, same header rules, same QC, same canonical
  output as the equivalent CSV;
* and does the reader refuse the workbook-specific cases where it would otherwise have to
  invent a value.
"""

from __future__ import annotations

from datetime import UTC, datetime, timedelta, timezone

import pytest
from openpyxl import Workbook

from virtualcell.core.experiment import MeasurementQuality, MeasurementValueType
from virtualcell.ingestion import (
    ColumnRole,
    ColumnSpec,
    DatasetSpec,
    IngestionStatus,
    QCRule,
    SourceFormat,
    ingest_file,
    ingest_table,
    read_delimited,
)
from virtualcell.ingestion.contracts import SPEC_VERSION, TimeAxisKind
from virtualcell.ingestion.readers import ReaderError, read_path
from virtualcell.ingestion.xlsx import read_workbook

HEADERS = ["cell_line", "passage", "PDL", "DT_min"]
ROWS = [["IMR 90", 25, 22.0, 2520], ["IMR 90", 30, 25.5, 4800], ["IMR 90", 35, None, 6000]]


def _spec(**over) -> DatasetSpec:
    fields = {
        "spec_version": SPEC_VERSION,
        "dataset_id": "fibroblast_passages",
        "source_format": SourceFormat.XLSX,
        "columns": [
            ColumnSpec(header="cell_line", role=ColumnRole.IDENTIFIER),
            ColumnSpec(header="passage", role=ColumnRole.TIME_AXIS, time_axis=TimeAxisKind.PASSAGE),
            ColumnSpec(
                header="PDL",
                role=ColumnRole.MEASUREMENT,
                name="cumulative_PDL",
                value_type=MeasurementValueType.NUMERIC,
                unit="population_doubling",
            ),
            ColumnSpec(
                header="DT_min",
                role=ColumnRole.MEASUREMENT,
                name="DT_hours",
                value_type=MeasurementValueType.NUMERIC,
                unit="hour",
                source_unit="minute",
                unit_factor=1 / 60,
            ),
        ],
        "conditions": {"medium": "DMEM"},
    }
    fields.update(over)
    return DatasetSpec(**fields)


def _workbook(tmp_path, rows=None, headers=None, name="passages.xlsx", sheets=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    for row in [headers or HEADERS, *(ROWS if rows is None else rows)]:
        sheet.append(row)
    for extra in sheets or []:
        workbook.create_sheet(extra)
    path = tmp_path / name
    workbook.save(path)
    return path


# --- the same contract as a delimited source ---------------------------------


def test_a_workbook_and_the_equivalent_csv_produce_the_same_run(tmp_path) -> None:
    """The point of putting xlsx behind the same DatasetSpec: the container changed, the
    meaning did not. Anything that differs here is the reader interpreting."""
    from virtualcell.core.experiment import dedup_key

    xlsx = ingest_file(_workbook(tmp_path), _spec())
    csv_text = (
        "cell_line,passage,PDL,DT_min\nIMR 90,25,22.0,2520\nIMR 90,30,25.5,4800\nIMR 90,35,,6000\n"
    )
    csv = ingest_table(
        read_delimited(csv_text, source_name="passages.xlsx", source_format=SourceFormat.CSV),
        _spec(source_format=SourceFormat.CSV),
    )

    assert xlsx.status is csv.status is IngestionStatus.SUCCESS
    # Identity is over observed content, so it holds across containers.
    assert dedup_key(xlsx.runs[0]) == dedup_key(csv.runs[0])
    assert xlsx.qc.counts == csv.qc.counts


def test_declared_normalization_and_qc_apply_unchanged(tmp_path) -> None:
    result = ingest_file(_workbook(tmp_path), _spec())
    assert result.status is IngestionStatus.SUCCESS

    first = result.runs[0].observations[0]
    doubling = next(m for m in first.measurements if m.name == "DT_hours")
    assert doubling.value == 42.0  # 2520 minutes, by the declared factor
    assert doubling.provenance.metadata["pre_normalization_value"] == 2520.0

    blank = next(
        m
        for o in result.runs[0].observations
        for m in o.measurements
        if m.quality is MeasurementQuality.MISSING
    )
    assert blank.value is None


def test_the_header_contract_is_the_shared_one(tmp_path) -> None:
    """Not re-implemented per format: the round-4 rules are enforced in ``build_table``,
    which every reader funnels through."""
    duplicate = _workbook(tmp_path, headers=["id", "id", "PDL", "DT_min"], name="dup.xlsx")
    with pytest.raises(ReaderError, match="duplicate header"):
        read_workbook(duplicate)

    empty = _workbook(tmp_path, headers=["cell_line", "", "PDL", "DT_min"], name="blank.xlsx")
    with pytest.raises(ReaderError, match="are empty"):
        read_workbook(empty)


def test_provenance_records_the_workbook_as_the_source(tmp_path) -> None:
    result = ingest_file(_workbook(tmp_path), _spec())
    metadata = result.runs[0].observations[0].measurements[0].provenance.metadata
    assert metadata["source_name"] == "passages.xlsx"
    assert metadata["column_header"] == "PDL"
    assert metadata["row_index"] == 0


# --- cell values are read as stored, not as displayed ------------------------


def test_an_integer_stored_as_a_float_is_not_made_fractional(tmp_path) -> None:
    """Excel stores 25 as 25.0. Rendering "25.0" would make a passage count look fractional
    to a reader that is right to refuse fractional passages."""
    path = _workbook(tmp_path, rows=[["IMR 90", 25.0, 22.0, 2520.0]], name="floats.xlsx")
    result = ingest_file(path, _spec())
    assert result.status is IngestionStatus.SUCCESS
    assert result.runs[0].observations[0].time_point.value == 25


def test_booleans_and_dates_are_rendered_in_forms_the_grammar_reads(tmp_path) -> None:
    columns = [
        *_spec().columns,
        ColumnSpec(
            header="contaminated",
            role=ColumnRole.MEASUREMENT,
            value_type=MeasurementValueType.BOOLEAN,
        ),
    ]
    path = _workbook(
        tmp_path,
        headers=[*HEADERS, "contaminated"],
        rows=[["IMR 90", 25, 22.0, 2520, False]],
        name="bool.xlsx",
    )
    result = ingest_file(path, _spec(columns=columns))
    flag = next(m for m in result.runs[0].observations[0].measurements if m.name == "contaminated")
    assert flag.value is False


def _timestamp_columns(offset: str | None):
    columns = [c for c in _spec().columns if c.header != "passage"]
    columns.append(
        ColumnSpec(
            header="passage",
            role=ColumnRole.TIME_AXIS,
            time_axis=TimeAxisKind.TIMESTAMP,
            timestamp_offset=offset,
        )
    )
    return columns


def test_excel_cannot_store_a_timezone_at_all(tmp_path) -> None:
    """Not a limitation of this reader: the format itself refuses. That is why a timestamp
    axis needs a *declared* offset to be usable from a spreadsheet at all."""
    workbook = Workbook()
    workbook.active.append(HEADERS)
    workbook.active.append(["IMR 90", datetime(2026, 1, 1, 12, 0, tzinfo=UTC), 22.0, 2520])
    with pytest.raises(TypeError, match="does not support timezones"):
        workbook.save(tmp_path / "impossible.xlsx")


def test_a_naive_timestamp_is_refused_when_no_offset_is_declared(tmp_path) -> None:
    """A naive stamp is ambiguous across sites and instruments, and the reader must not
    paper over that by inventing a zone."""
    path = _workbook(
        tmp_path, rows=[["IMR 90", datetime(2026, 1, 1, 12, 0), 22.0, 2520]], name="stamp.xlsx"
    )
    result = ingest_file(path, _spec(columns=_timestamp_columns(None)))
    assert result.status is IngestionStatus.NO_VALID_ROWS
    assert "timestamp_offset" in result.rejected_rows[0].detail


def test_a_declared_offset_makes_a_workbook_timestamp_usable(tmp_path) -> None:
    """Stated by a human, never inferred — the same rule as every other reading decision
    in this layer."""
    path = _workbook(
        tmp_path, rows=[["IMR 90", datetime(2026, 1, 1, 12, 0), 22.0, 2520]], name="stamp2.xlsx"
    )
    result = ingest_file(path, _spec(columns=_timestamp_columns("+09:00")))
    assert result.status is IngestionStatus.SUCCESS
    assert result.runs[0].observations[0].time_point.value == datetime(
        2026, 1, 1, 12, 0, tzinfo=timezone(timedelta(hours=9))
    )


def test_an_offset_declared_on_a_non_timestamp_axis_is_refused() -> None:
    with pytest.raises(ValueError, match="has no timezone"):
        ColumnSpec(
            header="p",
            role=ColumnRole.TIME_AXIS,
            time_axis=TimeAxisKind.PASSAGE,
            timestamp_offset="+09:00",
        )


def test_a_malformed_offset_is_refused() -> None:
    with pytest.raises(ValueError, match="ISO 8601 UTC offset"):
        ColumnSpec(
            header="t",
            role=ColumnRole.TIME_AXIS,
            time_axis=TimeAxisKind.TIMESTAMP,
            timestamp_offset="KST",
        )


def test_text_in_a_numeric_column_is_still_a_qc_failure(tmp_path) -> None:
    path = _workbook(tmp_path, rows=[["IMR 90", 25, "high", 2520]], name="text.xlsx")
    result = ingest_file(path, _spec())
    decision = next(d for d in result.qc.decisions if d.column == "cumulative_PDL")
    assert decision.rule is QCRule.UNPARSEABLE


# --- what the reader refuses rather than inventing ---------------------------


def test_a_multi_sheet_workbook_must_name_its_sheet(tmp_path) -> None:
    """Picking the first sheet is a guess about which experiment the file is about, and the
    wrong guess still imports cleanly."""
    path = _workbook(tmp_path, sheets=["notes"], name="multi.xlsx")
    result = ingest_file(path, _spec())
    assert result.status is IngestionStatus.UNREADABLE_SOURCE
    assert "names none" in result.errors[0]

    assert ingest_file(path, _spec(sheet="data")).status is IngestionStatus.SUCCESS


def test_a_named_sheet_that_does_not_exist_is_refused(tmp_path) -> None:
    path = _workbook(tmp_path, sheets=["notes"], name="multi2.xlsx")
    result = ingest_file(path, _spec(sheet="absent"))
    assert result.status is IngestionStatus.UNREADABLE_SOURCE
    assert "no sheet named" in result.errors[0]


def test_a_formula_with_no_cached_value_is_refused(tmp_path) -> None:
    """openpyxl does not evaluate formulas; it reads what Excel last saved. With no cached
    value the cell reads blank, which QC would record as *missing* — asserting no reading
    was taken when one exists and cannot be seen. That is a lie about the data."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["IMR 90", 25, "=B2*2", 2520])
    path = tmp_path / "formula.xlsx"
    workbook.save(path)

    result = ingest_file(path, _spec())
    assert result.status is IngestionStatus.UNREADABLE_SOURCE
    assert "no cached value" in result.errors[0]


def test_a_merged_cell_is_refused(tmp_path) -> None:
    """A merged block has one value and several coordinates; a locator naming one row and
    one column cannot describe it, and the non-anchor cells read as blank."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["IMR 90", 25, 22.0, 2520])
    sheet.merge_cells("C2:D2")
    path = tmp_path / "merged.xlsx"
    workbook.save(path)

    result = ingest_file(path, _spec())
    assert result.status is IngestionStatus.UNREADABLE_SOURCE
    assert "merged range" in result.errors[0]


def test_an_excel_error_value_is_refused(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["IMR 90", 25, "#DIV/0!", 2520])
    path = tmp_path / "error.xlsx"
    workbook.save(path)

    result = ingest_file(path, _spec())
    assert result.status is IngestionStatus.UNREADABLE_SOURCE
    assert "error value" in result.errors[0]


def test_a_file_that_is_not_a_workbook_is_a_typed_status(tmp_path) -> None:
    path = tmp_path / "not.xlsx"
    path.write_text("cell_line,passage\nIMR 90,25\n", encoding="utf-8")
    result = ingest_file(path, _spec())
    assert result.status is IngestionStatus.UNREADABLE_SOURCE
    assert result.errors


def test_a_missing_workbook_is_a_typed_status(tmp_path) -> None:
    result = ingest_file(tmp_path / "absent.xlsx", _spec())
    assert result.status is IngestionStatus.UNREADABLE_SOURCE


# --- the spec field ----------------------------------------------------------


def test_sheet_is_meaningful_only_for_a_workbook() -> None:
    """A delimited file has no worksheets, so declaring one signals a spec that does not
    describe the file it is pointed at."""
    with pytest.raises(ValueError, match="only for an xlsx source"):
        _spec(source_format=SourceFormat.CSV, sheet="data")


def test_a_spec_written_before_the_sheet_field_still_reads() -> None:
    """``sheet`` is additive, so 1.0 specs are unaffected — the minor bump exists to record
    that, not to invalidate them."""
    assert _spec(spec_version="1.0").sheet is None


def test_read_path_dispatches_on_the_declared_format(tmp_path) -> None:
    table = read_path(_workbook(tmp_path), source_format=SourceFormat.XLSX)
    assert table.headers == HEADERS
    assert SourceFormat.XLSX.is_delimited is False
    assert SourceFormat.CSV.is_delimited and SourceFormat.TSV.is_delimited


# --- review round 5: an offset must actually name a zone ---------------------


@pytest.mark.parametrize("offset", ["", "   "])
def test_an_offset_that_names_no_timezone_is_refused(offset: str) -> None:
    """The defect: an empty offset appends nothing, so the validation probe parsed cleanly
    and stayed *naive*. It then reached ingestion, left the stamp naive, and crashed inside
    TimestampTimePoint. Declaring an offset that names no zone is worse than declaring none,
    because it reads as an answer to exactly the ambiguity it fails to resolve."""
    with pytest.raises(ValueError, match="names no timezone|ISO 8601 UTC offset"):
        ColumnSpec(
            header="t",
            role=ColumnRole.TIME_AXIS,
            time_axis=TimeAxisKind.TIMESTAMP,
            timestamp_offset=offset,
        )


@pytest.mark.parametrize("offset", ["+09:00", "-05:00", "Z", "+00:00"])
def test_offsets_that_do_name_a_zone_are_accepted(offset: str) -> None:
    column = ColumnSpec(
        header="t",
        role=ColumnRole.TIME_AXIS,
        time_axis=TimeAxisKind.TIMESTAMP,
        timestamp_offset=offset,
    )
    assert column.timestamp_offset == offset


def test_a_naive_stamp_never_reaches_canonical_construction(tmp_path) -> None:
    """The second lock on the same door. Spec validation now refuses a zone-less offset, so
    this can only be reached by constructing around it — but if it ever is, the row is
    rejected cleanly instead of raising halfway through an import."""
    from virtualcell.ingestion.contracts import CellLocator, RawCell
    from virtualcell.ingestion.parse import parse_cell

    column = ColumnSpec(
        header="t",
        role=ColumnRole.TIME_AXIS,
        time_axis=TimeAxisKind.TIMESTAMP,
        timestamp_offset="+09:00",
    )
    # model_construct bypasses validation, standing in for any future path that skips it.
    zoneless = column.model_copy(update={"timestamp_offset": ""})
    cell = RawCell(
        locator=CellLocator(source_name="s.xlsx", row_index=0, column_header="t"),
        text="2026-01-01T12:00:00",
    )
    parsed = parse_cell(cell, zoneless, _spec())
    assert parsed.value is None
    assert "timezone-naive" in (parsed.parse_note or "")


def test_a_zoneless_offset_cannot_crash_an_xlsx_import(tmp_path) -> None:
    """End to end: the spec is refused before any workbook is opened."""
    path = _workbook(
        tmp_path, rows=[["IMR 90", datetime(2026, 1, 1, 12, 0), 22.0, 2520]], name="zoneless.xlsx"
    )
    with pytest.raises(ValueError, match="names no timezone"):
        _spec(columns=_timestamp_columns(""))
    # ...and the same file with a real offset imports cleanly.
    assert ingest_file(path, _spec(columns=_timestamp_columns("Z"))).status is (
        IngestionStatus.SUCCESS
    )
