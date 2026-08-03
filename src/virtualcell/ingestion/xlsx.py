"""XLSX reading (PR13b-2), behind the same :class:`DatasetSpec` as CSV and TSV.

A spreadsheet is not a delimited file with a different separator. It is a typed, formula-
bearing, multi-sheet container, and each of those is a way for a value to arrive different
from what a human saw in Excel. This module's job is to hand :func:`build_table` rows of
**text** that faithfully represent the cells, and to refuse the file when it cannot.

What it refuses, and why each would otherwise be silent:

* **An unnamed sheet in a multi-sheet workbook.** Picking the first is a guess about which
  experiment the file is about, and the wrong guess still imports cleanly.
* **A formula with no cached value.** ``openpyxl`` does not evaluate formulas; it reads the
  value Excel last saved. When there is none, the cell reads as blank — which QC would
  record as *missing*, asserting no reading was taken when in fact one exists and cannot be
  seen. That is a lie about the data, not a gap in it.
* **Merged cells over the used range.** A merged block has one value and several
  coordinates; the non-anchor coordinates read as blank. A :class:`CellLocator` names one
  row and one column, so a merged cell has no locator that means anything.
* **An error value** (``#DIV/0!``, ``#REF!``). Excel already knows the cell is wrong.

Deliberately not here: reading cell formatting, colours, comments, charts or defined names.
Presentation is not data, and a value that only exists in a cell's formatting is a value
the file did not really record.
"""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path

from virtualcell.ingestion.contracts import RawTable
from virtualcell.ingestion.readers import ReaderError, build_table

# Excel's in-cell error sentinels. They arrive as plain strings, so without this they would
# look like ordinary categorical text.
_ERROR_VALUES = frozenset(
    {"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!", "#SPILL!", "#CALC!"}
)

_INSTALL_HINT = (
    "reading .xlsx requires the optional 'xlsx' extra: pip install 'virtualcell[xlsx]'. "
    "It is optional because a spreadsheet parser is a dependency only some imports need; "
    "exporting the sheet to CSV is always an alternative."
)


def _load(path: Path):
    """Load a workbook twice: once for values, once to see which cells are formulas.

    ``data_only=True`` yields Excel's last cached result and ``None`` for a formula that was
    never evaluated — indistinguishable from a genuinely empty cell. Reading the formulas as
    well is the only way to tell those two apart, and telling them apart is the whole point:
    one is a missing reading, the other is a reading this reader cannot see.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised only without the extra
        raise ReaderError(f"cannot read {path.name}: {_INSTALL_HINT}") from exc

    if not path.exists():
        raise ReaderError(f"cannot read {path}: no such file")
    try:
        values = load_workbook(path, data_only=True, read_only=True)
        # Not read-only: openpyxl's optimised reader does not expose merged ranges, and a
        # merged block that goes unnoticed reads as blank — the silent failure this whole
        # module is arranged to prevent. Lab exports are small enough for the full loader.
        formulas = load_workbook(path, data_only=False, read_only=False)
    except OSError as exc:
        raise ReaderError(f"cannot read {path}: {exc}") from exc
    except Exception as exc:  # openpyxl raises a wide range for a malformed workbook
        raise ReaderError(
            f"cannot read {path.name} as an xlsx workbook: {type(exc).__name__}: {exc}"
        ) from exc
    return values, formulas


def _select(workbook, path: Path, sheet: str | None):
    """Pick the sheet to read, refusing to guess when a workbook holds several."""
    names = list(workbook.sheetnames)
    if sheet is not None:
        if sheet not in names:
            raise ReaderError(
                f"{path.name}: no sheet named {sheet!r}; the workbook contains "
                f"{', '.join(repr(n) for n in names)}"
            )
        return workbook[sheet]
    if len(names) > 1:
        raise ReaderError(
            f"{path.name}: the workbook has {len(names)} sheets "
            f"({', '.join(repr(n) for n in names)}) and the spec names none. Declare "
            "'sheet' — choosing one for you would be a guess about which experiment this "
            "file is about, and the wrong guess still imports cleanly"
        )
    return workbook[names[0]]


def _as_text(value: object) -> str:
    """Render one cell as the text the rest of the pipeline parses.

    Numbers are rendered with ``repr`` semantics rather than Excel's *display* formatting:
    the stored value is the datum, and a column formatted to two decimals has not actually
    rounded anything. Dates become ISO 8601 so the canonical timestamp axis can read them —
    Excel stores no timezone, so a bare stamp is refused later by the axis parser, which is
    the correct place for that judgement.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        # 25.0 is how Excel stores the integer 25; rendering "25.0" would make a passage
        # count look fractional to a reader that is right to be strict about it.
        return str(int(value))
    return str(value)


def _merged_coordinates(sheet) -> set[tuple[int, int]]:
    ranges = getattr(sheet, "merged_cells", None)
    if ranges is None:
        return set()
    coordinates: set[tuple[int, int]] = set()
    for merged in getattr(ranges, "ranges", []):
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                coordinates.add((row, column))
    return coordinates


def read_workbook(path: str | Path, *, sheet: str | None = None) -> RawTable:
    """Read one worksheet into a :class:`RawTable`, verbatim.

    The header contract (unique, non-empty) is enforced by :func:`build_table`, shared with
    the delimited readers, so an xlsx file and a csv file are held to the same rules.
    """
    file_path = Path(path)
    values, formulas = _load(file_path)
    try:
        value_sheet = _select(values, file_path, sheet)
        formula_sheet = _select(formulas, file_path, sheet)

        # A merged range is read from the workbook that was *not* opened read-only-optimised
        # for values; openpyxl exposes merged ranges on the formula view reliably.
        merged = _merged_coordinates(formula_sheet)

        rows: list[list[str]] = []
        unevaluated: list[str] = []
        errors: list[str] = []

        value_rows = value_sheet.iter_rows(values_only=False)
        formula_rows = formula_sheet.iter_rows(values_only=True)
        for row_index, (value_row, formula_row) in enumerate(
            zip(value_rows, formula_rows, strict=False), start=1
        ):
            texts: list[str] = []
            for column_index, cell in enumerate(value_row, start=1):
                if (row_index, column_index) in merged:
                    raise ReaderError(
                        f"{file_path.name}: cell {cell.coordinate} is part of a merged "
                        "range. A merged block has one value and several coordinates, so a "
                        "locator naming one row and one column cannot describe it; unmerge "
                        "the range or export the sheet to CSV"
                    )
                raw = cell.value
                formula = (
                    formula_row[column_index - 1] if column_index <= len(formula_row) else None
                )
                if isinstance(raw, str) and raw in _ERROR_VALUES:
                    errors.append(f"{cell.coordinate}={raw}")
                if raw is None and isinstance(formula, str) and formula.startswith("="):
                    unevaluated.append(cell.coordinate)
                texts.append(_as_text(raw))
            rows.append(texts)

        if errors:
            raise ReaderError(
                f"{file_path.name}: {len(errors)} cell(s) hold an Excel error value "
                f"({', '.join(errors[:5])}); the spreadsheet already knows these are wrong"
            )
        if unevaluated:
            raise ReaderError(
                f"{file_path.name}: {len(unevaluated)} formula cell(s) have no cached value "
                f"({', '.join(unevaluated[:5])}). This reader does not evaluate formulas, "
                "and reading them as blank would record 'no reading was taken' for cells "
                "that do hold one; open and re-save the workbook, or export it to CSV"
            )
    finally:
        values.close()
        formulas.close()

    # Trailing all-empty columns are an artifact of Excel's used range, not data. They are
    # trimmed *before* the header contract runs, so a stray formatted column does not read
    # as an unnamed one.
    while rows and rows[0] and not rows[0][-1].strip():
        if any(row[-1].strip() for row in rows if row):
            break
        for row in rows:
            if row:
                row.pop()

    return build_table(rows, source_name=file_path.name)
